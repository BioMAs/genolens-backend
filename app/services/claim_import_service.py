"""
Cosmetics module — claim referential import service.

Parses the SciLicium "Claims_Pathway_Review_*.xlsx" workbooks into the
claim_pathway_mappings / claim_references tables, and seeds the canonical
cosmetic_claims taxonomy. Admin-only (callers must enforce authorization).

The workbooks ship with the backend at app/data/claims/. They can also be
re-imported via an admin upload endpoint (import_workbook_bytes).
"""
from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Optional, Union

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.models import (
    ClaimDirection,
    ClaimPathwayMapping,
    ClaimReference,
    CosmeticClaim,
    EvidenceLevel,
)
from app.services.cosmetics_taxonomy import (
    CANONICAL_CLAIMS,
    derive_canonical_claims,
    normalize_pathway_id,
)

logger = logging.getLogger(__name__)

# Default location of the bundled referential workbooks.
DEFAULT_CLAIMS_DIR = Path(__file__).resolve().parent.parent / "data" / "claims"

_PATHWAYS_SHEET = "Pathways"
_REFERENCES_SHEET = "References"

# Header -> model attribute for the Pathways sheet.
_PATHWAY_COLS = {
    "Term_ID": "term_id",
    "Description": "description",
    "Common_abbrev": "common_abbrev",
    "Suggested_acronym": "suggested_acronym",
    "Original_claims": "original_claims",
    "Original_direction": "original_direction",
    "Updated_claim_framing": "updated_claim_framing",
    "Updated_direction": "updated_direction",
    "Category": "category",
    "Evidence_level": "evidence_level",
    "Rationale": "rationale",
    "Caveats": "caveats",
    "RefCat": "ref_cat",
}


def _clean(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _coerce_direction(value: Optional[str]) -> ClaimDirection:
    if not value:
        return ClaimDirection.UNKNOWN
    try:
        return ClaimDirection(value.strip().upper())
    except ValueError:
        return ClaimDirection.UNKNOWN


def _coerce_evidence(value: Optional[str]) -> EvidenceLevel:
    if not value:
        return EvidenceLevel.LOW
    try:
        return EvidenceLevel(value.strip().upper())
    except ValueError:
        return EvidenceLevel.LOW


def _parse_workbook(source: Union[Path, bytes]) -> tuple[list[dict], list[dict]]:
    """Return (pathway_rows, reference_rows) parsed from one workbook."""
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)

    pathway_rows: list[dict] = []
    reference_rows: list[dict] = []

    if _PATHWAYS_SHEET in wb.sheetnames:
        ws = wb[_PATHWAYS_SHEET]
        rows = ws.iter_rows(values_only=True)
        header = [(_clean(h) or "") for h in next(rows, [])]
        idx = {name: header.index(name) for name in _PATHWAY_COLS if name in header}
        for row in rows:
            term_id = _clean(row[idx["Term_ID"]]) if "Term_ID" in idx else None
            if not term_id:
                continue
            rec = {attr: _clean(row[idx[col]]) for col, attr in _PATHWAY_COLS.items() if col in idx}
            pathway_rows.append(rec)

    if _REFERENCES_SHEET in wb.sheetnames:
        ws = wb[_REFERENCES_SHEET]
        rows = ws.iter_rows(values_only=True)
        header = [(_clean(h) or "") for h in next(rows, [])]
        try:
            ci_cat = header.index("RefCat")
            ci_sum = header.index("Source_summary")
            ci_url = header.index("URL")
        except ValueError:
            ci_cat = ci_sum = ci_url = None
        if ci_cat is not None:
            for row in rows:
                ref_cat = _clean(row[ci_cat])
                if not ref_cat:
                    continue
                reference_rows.append(
                    {
                        "ref_cat": ref_cat,
                        "source_summary": _clean(row[ci_sum]) if ci_sum is not None else None,
                        "url": _clean(row[ci_url]) if ci_url is not None else None,
                    }
                )

    wb.close()
    return pathway_rows, reference_rows


async def _upsert_mappings(db: AsyncSession, rows: list[dict]) -> tuple[int, int]:
    """Upsert pathway rows keyed by term_id. Returns (created, updated)."""
    created = updated = 0
    existing = {
        m.term_id: m
        for m in (await db.execute(select(ClaimPathwayMapping))).scalars().all()
    }
    for rec in rows:
        term_id = rec["term_id"]
        direction = _coerce_direction(rec.get("updated_direction"))
        evidence = _coerce_evidence(rec.get("evidence_level"))
        canonical = derive_canonical_claims(rec.get("original_claims"), rec.get("category"))
        payload = dict(
            term_id=term_id,
            term_id_normalized=normalize_pathway_id(term_id),
            description=rec.get("description"),
            common_abbrev=rec.get("common_abbrev"),
            suggested_acronym=rec.get("suggested_acronym"),
            original_claims=rec.get("original_claims"),
            original_direction=rec.get("original_direction"),
            updated_claim_framing=rec.get("updated_claim_framing"),
            updated_direction=direction,
            category=rec.get("category"),
            evidence_level=evidence,
            rationale=rec.get("rationale"),
            caveats=rec.get("caveats"),
            ref_cat=rec.get("ref_cat"),
            canonical_claims=canonical,
            is_active=True,
        )
        obj = existing.get(term_id)
        if obj is None:
            obj = ClaimPathwayMapping(**payload)
            db.add(obj)
            existing[term_id] = obj
            created += 1
        else:
            for k, v in payload.items():
                setattr(obj, k, v)
            updated += 1
    return created, updated


async def _upsert_references(db: AsyncSession, rows: list[dict]) -> int:
    """Insert reference rows, de-duplicated by (ref_cat, url)."""
    existing = {
        (r.ref_cat, r.url)
        for r in (await db.execute(select(ClaimReference))).scalars().all()
    }
    added = 0
    for rec in rows:
        key = (rec["ref_cat"], rec.get("url"))
        if key in existing:
            continue
        db.add(ClaimReference(**rec))
        existing.add(key)
        added += 1
    return added


async def seed_cosmetic_claims(db: AsyncSession) -> int:
    """Upsert the canonical claim taxonomy. Returns number of rows touched."""
    existing = {
        c.slug: c for c in (await db.execute(select(CosmeticClaim))).scalars().all()
    }
    touched = 0
    for spec in CANONICAL_CLAIMS:
        obj = existing.get(spec["slug"])
        if obj is None:
            db.add(CosmeticClaim(**spec, is_active=True))
        else:
            for k, v in spec.items():
                setattr(obj, k, v)
            obj.is_active = True
        touched += 1
    return touched


async def import_claim_workbooks(
    db: AsyncSession, paths: Optional[list[Path]] = None
) -> dict:
    """Import every bundled workbook (or the given paths) and seed the taxonomy."""
    if paths is None:
        paths = sorted(DEFAULT_CLAIMS_DIR.glob("Claims_Pathway_Review_*.xlsx"))
    if not paths:
        raise FileNotFoundError(f"No claim workbooks found in {DEFAULT_CLAIMS_DIR}")

    all_mappings: list[dict] = []
    all_refs: list[dict] = []
    for p in paths:
        m, r = _parse_workbook(p)
        all_mappings.extend(m)
        all_refs.extend(r)

    created, updated = await _upsert_mappings(db, all_mappings)
    refs_added = await _upsert_references(db, all_refs)
    claims_seeded = await seed_cosmetic_claims(db)
    await db.commit()

    result = {
        "files": len(paths),
        "mappings_created": created,
        "mappings_updated": updated,
        "references_added": refs_added,
        "claims_seeded": claims_seeded,
    }
    logger.info("Claim referential import: %s", result)
    return result


async def import_workbook_bytes(db: AsyncSession, data: bytes) -> dict:
    """Import a single uploaded workbook (admin upload endpoint)."""
    mappings, refs = _parse_workbook(data)
    created, updated = await _upsert_mappings(db, mappings)
    refs_added = await _upsert_references(db, refs)
    claims_seeded = await seed_cosmetic_claims(db)
    await db.commit()
    return {
        "files": 1,
        "mappings_created": created,
        "mappings_updated": updated,
        "references_added": refs_added,
        "claims_seeded": claims_seeded,
    }
